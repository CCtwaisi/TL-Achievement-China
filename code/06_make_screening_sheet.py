#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Make Title/Abstract screening Excel from master_refs_dedup/raw

import os, hashlib, pandas as pd

def read_any(path):
    # 自动尝试多种编码，避免中文乱码/报错
    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return pd.read_csv(path, dtype=str, encoding=enc).fillna("")
        except Exception:
            pass
    # 最后再试一次，让报错抛出
    return pd.read_csv(path, dtype=str)

os.makedirs("data_clean", exist_ok=True)

src = None
for p in ["data_clean/master_refs_dedup.csv", "data_clean/master_refs_raw.csv"]:
    if os.path.exists(p):
        src = p
        break
if src is None:
    raise SystemExit("No input CSV found. Put data_clean/master_refs_dedup.csv or master_refs_raw.csv")

df = read_any(src)
cols = [c.lower() for c in df.columns]
def pick(*names):
    for n in names:
        if n.lower() in cols: return n

col_title   = pick("title","ti")
col_authors = pick("authors","au","author")
col_year    = pick("year","py","yr")
col_journal = pick("journal","source","so")
col_abs     = pick("abstract","ab")
col_kw      = pick("keywords","kw")
col_doi     = pick("doi")
col_url     = pick("url","link")
use = [c for c in [col_title,col_authors,col_year,col_journal,col_abs,col_kw,col_doi,col_url] if c]
work = df[use].copy() if use else df.copy()

# study_id
if "study_id" in df.columns:
    work.insert(0,"study_id",df["study_id"].astype(str))
else:
    ids = [(hashlib.md5(((r.get(col_title,"") or "")+"|"+(r.get(col_year,"") or "")).encode("utf-8","ignore")).hexdigest()[:12]) for _,r in df.iterrows()]
    work.insert(0,"study_id",ids)

# Screening columns
work.insert(1,"incl_titleabs_yesno","")
work.insert(2,"exclusion_reason","")

# 写 Excel
out_xlsx = "data_clean/TitleAbs_Screening.xlsx"
with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
    work.to_excel(w, index=False, sheet_name="Screening")
    pd.DataFrame({"Allowed_Values":["NOT_MIDDLE_SCHOOL","NOT_CHINA","NOT_QUANT","NO_EFFECT_SIZE","DUPLICATE_DATASET","THEORY_ONLY","OTHER"]}).to_excel(w, index=False, sheet_name="Dictionary")
    pd.DataFrame({"YesNo":["yes","no"]}).to_excel(w, index=False, sheet_name="YesNo")

# 下拉菜单
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
wb = load_workbook(out_xlsx); ws = wb["Screening"]; max_row = ws.max_row
dv_yesno = DataValidation(type="list", formula1="=YesNo!$A$2:$A$3", allow_blank=True); ws.add_data_validation(dv_yesno); dv_yesno.add(f"B2:B{max_row}")
last_reason = wb["Dictionary"].max_row
dv_reason = DataValidation(type="list", formula1=f"=Dictionary!$A$2:$A${last_reason}", allow_blank=True); ws.add_data_validation(dv_reason); dv_reason.add(f"C2:C{max_row}")
wb.save(out_xlsx)
print(f"Created {out_xlsx}")
